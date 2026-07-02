"""生成 Markdown + xlsx 报告。"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from .models import StudentResult


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def write_markdown(results: list[StudentResult], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"recommendations_{_timestamp()}.md"

    lines = ["# 课题推荐结果", "", f"生成时间：{datetime.now():%Y-%m-%d %H:%M:%S}", ""]
    review_needed = []

    for r in results:
        lines.append(f"## {r.student.name}（投递岗位：{r.student.position or '未填'}）")
        lines.append("")
        if r.error:
            lines.append(f"> ⚠️ 推荐失败，需人工复核：{r.error}")
            lines.append("")
            review_needed.append(r.student.name)
            continue
        lines.append("| 排名 | 课题编号 | 课题 | 匹配度 | 推荐理由 |")
        lines.append("| --- | --- | --- | --- | --- |")
        for rank, item in enumerate(r.items, start=1):
            reason = item.reason.replace("\n", " ").replace("|", "／")
            lines.append(
                f"| {rank} | {item.topic_id} | {item.topic_name} | {item.score:.0f} | {reason} |"
            )
        lines.append("")

    if review_needed:
        lines.append("---")
        lines.append("")
        lines.append(f"### ⚠️ 待人工复核（{len(review_needed)} 人）")
        lines.append("")
        lines.append("、".join(review_needed))
        lines.append("")

    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def write_xlsx(results: list[StudentResult], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"recommendations_{_timestamp()}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "课题推荐"

    header = ["姓名", "投递岗位", "排名", "课题编号", "课题", "匹配度", "推荐理由", "备注"]
    ws.append(header)
    head_fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in results:
        if r.error:
            ws.append([r.student.name, r.student.position, "", "", "", "", "", f"复核：{r.error}"])
            continue
        for rank, item in enumerate(r.items, start=1):
            ws.append([
                r.student.name,
                r.student.position,
                rank,
                item.topic_id,
                item.topic_name,
                round(item.score),
                item.reason,
                "",
            ])

    # 列宽 + 理由列自动换行
    widths = [12, 16, 6, 10, 24, 8, 60, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        row[6].alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(path)
    return path
